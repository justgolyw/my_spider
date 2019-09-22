#!/usr/bin/env python
# -*- coding:utf-8 -*-

import requests, bs4, time, random, socket, http.client, csv
import openpyxl,xlwt


def getContent(url, data=None):
    # 消息请求头
    header = {
        'Accept': 'text/html,*/*,q=0.1',
        'Accept-Encoding': 'gzip,deflate,sdch',
        'Accept-language': 'zh-CN,zh;q=0.8',
        'Connection': 'keep-alive',
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) Apple'
    }

    # 请求url地址，获得返回 response 信息
    response = requests.get(url, headers=header)
    response.encoding = 'utf-8'
    print('request success')
    return response.text  # 返回Html 全文


def getData(html):
    final = []
    bs = bs4.BeautifulSoup(html, features='html.parser')
    body = bs.body  # 获取body
    data = body.find('div', {'id': '7d'})
    ul = data.find('ul')
    li = ul.find_all('li')

    for day in li:
        temp = []
        date = day.find('h1').text  # 日期
        temp.append(date)
        inf = day.find_all('p')
        weather = inf[0].text  # 天气
        temp.append(weather)
        temperature_high = inf[1].find('span').text  # 最高温度
        temp.append(temperature_high)
        temperature_low = inf[1].find('i').text  # 最低温度
        temp.append(temperature_low)
        final.append(temp)
    print('getDate success')
    print(final)
    return final




# 将数据写入csv
def writeData(data, filename):
    with open(filename, mode='w', errors='ignore') as f:
        f_csv = csv.writer(f, delimiter=',', lineterminator='\n\n')
        f_csv.writerows(data)
    print('write_csv success')


# 将数据写入Excel保存
def writeToExcel(data, filename):
    # 新建excel文件
    wb = openpyxl.Workbook()
    # 新建sheet
    sheet = wb.active
    # sheet 命名
    sheet.title = 'weather'
    sheet['A1'] = '日期'
    sheet['B1'] = '天气'
    sheet['C1'] = '最高温度'
    sheet['D1'] = '最低温度'
    # index从1开始
    for i in range(2, 2 + len(data)):
        for j in range(1, 1 + len(data[0])):
            sheet.cell(row=i, column=j).value = data[i - 2][j - 1]
    wb.save(filename)


def writeToExcel2(data,filename):
    # 新建excel
    file = xlwt.Workbook()
    # 新建sheet
    sheet = file.add_sheet('weather',cell_overwrite_ok=True)
    # 表头命名
    #index从0开始
    sheet.write(0, 0, '日期')
    sheet.write(0, 1, '天气')
    sheet.write(0, 2, '最高温度')
    sheet.write(0, 3, '最低温度')
    # 写入数据
    # 设置样式

    style = xlwt.XFStyle()
    style2 = xlwt.easyxf('font: name Times New Roman, color-index red, bold on',
               num_format_str='#,##0.00')
    font = xlwt.Font()

    font.name = 'Times New Roman'
    font.bold = True
    style.font = font
    for i in range(1, 1 + len(data)): # 行
        for j in range(0, len(data[0])): # 列
            sheet.write(i,j,data[i-1][j],style2)

    file.save(filename)


if __name__ == '__main__':
    url = 'http://www.weather.com.cn/weather/101020100.shtml'
    html = getContent(url)  # 获取网页信息
    data = getData(html)  # 获取数据
    # writeToExcel(data, 'weather.xlsx')

    writeToExcel2(data,'example.xls')# 注意后缀名为xls不是xlsx
